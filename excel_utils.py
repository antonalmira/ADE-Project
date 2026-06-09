import openpyxl
import re


def extract_excel_table(worksheet):
    """Extract and filter table data from an Excel worksheet, including merged cells."""
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    if max_row == 0 or max_col == 0:
        print("Worksheet is empty, returning empty table")
        return [], []

    table_data = []
    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        row_data = []
        for cell in row:
            cell_value = cell.value
            if isinstance(cell_value, (int, float)) and cell_value is not None:
                cell_value = f"{cell_value:.3g}"
            else:
                cell_value = str(cell_value) if cell_value is not None else ""
            row_data.append(cell_value)
        table_data.append(row_data)

    filtered_rows = []
    row_indices_to_keep = []
    for row_idx, row in enumerate(table_data):
        if any(cell and str(cell).strip() for cell in row):
            filtered_rows.append(row)
            row_indices_to_keep.append(row_idx)

    if not filtered_rows:
        print("All rows are empty after filtering, returning empty table")
        return [], []

    transposed_data = list(map(list, zip(*filtered_rows)))
    filtered_columns = []
    col_indices_to_keep = []
    for col_idx, col in enumerate(transposed_data):
        if any(cell and str(cell).strip() for cell in col):
            filtered_columns.append(col)
            col_indices_to_keep.append(col_idx)

    if not filtered_columns:
        print("All columns are empty after filtering, returning empty table")
        return [], []

    filtered_table = list(map(list, zip(*filtered_columns)))

    merged_cells = []
    for merged_range in worksheet.merged_cells.ranges:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row - 1,
            merged_range.min_col - 1,
            merged_range.max_row - 1,
            merged_range.max_col - 1
        )
        row_intersects = any(min_row <= row_idx <= max_row for row_idx in row_indices_to_keep)
        col_intersects = any(min_col <= col_idx <= max_col for col_idx in col_indices_to_keep)
        if row_intersects and col_intersects:
            new_min_row = sum(1 for r in row_indices_to_keep if r < min_row)
            new_max_row = sum(1 for r in row_indices_to_keep if r <= max_row) - 1
            new_min_col = sum(1 for c in col_indices_to_keep if c < min_col)
            new_max_col = sum(1 for c in col_indices_to_keep if c <= max_col) - 1
            if new_min_row <= new_max_row and new_min_col <= new_max_col:
                merged_cells.append((new_min_row, new_min_col, new_max_row, new_max_col))

    print(f"Filtered table: {len(filtered_table)} rows, {len(filtered_table[0]) if filtered_table else 0} columns")
    print(f"Adjusted merged cells: {merged_cells}")
    return filtered_table, merged_cells


def peek_table_voltages(file_path, sheet_name):
    """Smarter function to quickly peek into an Excel table and find distinct voltages."""
    wb = None
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return []

        ws = wb[sheet_name]
        raw_data = [list(row) for row in ws.iter_rows(values_only=True, max_row=200)]

        group_col_idx = -1

        # 1. Look for the column by checking headers first
        header_keywords = ['vac', 'vin', 'v_in', 'input', 'input voltage', 'input (vac)', 'line voltage']
        for i, row in enumerate(raw_data[:10]):
            for j, cell in enumerate(row):
                val = str(cell).strip().lower()
                if val in header_keywords or 'input voltage' in val or 'input (vac)' in val:
                    group_col_idx = j
                    break
            if group_col_idx != -1:
                break

        # 2. Fallback: use safe hardcoded voltage values to identify the column
        if group_col_idx == -1:
            safe_pattern = re.compile(
                r'^(85|90|100|115|132|180|230|264|265|277)(\.0)?\s*(vac|v)?$', re.IGNORECASE
            )
            for i, row in enumerate(raw_data):
                for j, cell in enumerate(row):
                    val = str(cell).strip() if cell is not None else ""
                    if safe_pattern.match(val):
                        group_col_idx = j
                        break
                if group_col_idx != -1:
                    break

        if group_col_idx == -1:
            return []

        # 3. Extract any voltage value from that specific column
        generic_pattern = re.compile(r'^(\d{2,3}(\.\d+)?)\s*(vac|v)?$', re.IGNORECASE)

        voltages = []
        for i in range(len(raw_data)):
            val = str(raw_data[i][group_col_idx]).strip() if raw_data[i][group_col_idx] is not None else ""
            match = generic_pattern.match(val)
            if match:
                v = match.group(1)
                if v not in voltages:
                    voltages.append(v)

        return voltages

    except Exception as e:
        print(f"Error peeking table {sheet_name}: {e}")
        return []

    finally:
        # Always close the workbook to release the file lock, even if an exception occurred
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass