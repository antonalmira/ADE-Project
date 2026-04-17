import openpyxl

def extract_excel_table(worksheet):
    """Extract and filter table data from an Excel worksheet, including merged cells."""
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    if max_row == 0 or max_col == 0:
        print("Worksheet is empty, returning empty table")
        return [], []

    table_data = []
    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        row_data = []
        for cell in row:
            cell_value = cell.value
            if isinstance(cell_value, (int, float)) and cell_value is not None:
                cell_value = f"{cell_value:.3g}"
            else:
                cell_value = str(cell_value) if cell_value is not None else ""
            row_data.append(cell_value)
        table_data.append(row_data)

    filtered_rows = []
    row_indices_to_keep = []
    for row_idx, row in enumerate(table_data):
        if any(cell and str(cell).strip() for cell in row):
            filtered_rows.append(row)
            row_indices_to_keep.append(row_idx)

    if not filtered_rows:
        print("All rows are empty after filtering, returning empty table")
        return [], []

    transposed_data = list(map(list, zip(*filtered_rows)))
    filtered_columns = []
    col_indices_to_keep = []
    for col_idx, col in enumerate(transposed_data):
        if any(cell and str(cell).strip() for cell in col):
            filtered_columns.append(col)
            col_indices_to_keep.append(col_idx)

    if not filtered_columns:
        print("All columns are empty after filtering, returning empty table")
        return [], []

    filtered_table = list(map(list, zip(*filtered_columns)))

    merged_cells = []
    for merged_range in worksheet.merged_cells.ranges:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row - 1,
            merged_range.min_col - 1,
            merged_range.max_row - 1,
            merged_range.max_col - 1
        )
        row_intersects = any(min_row <= row_idx <= max_row for row_idx in row_indices_to_keep)
        col_intersects = any(min_col <= col_idx <= max_col for col_idx in col_indices_to_keep)
        if row_intersects and col_intersects:
            new_min_row = sum(1 for r in row_indices_to_keep if r < min_row)
            new_max_row = sum(1 for r in row_indices_to_keep if r <= max_row) - 1
            new_min_col = sum(1 for c in col_indices_to_keep if c < min_col)
            new_max_col = sum(1 for c in col_indices_to_keep if c <= max_col) - 1
            if new_min_row <= new_max_row and new_min_col <= new_max_col:
                merged_cells.append((new_min_row, new_min_col, new_max_row, new_max_col))

    print(f"Filtered table: {len(filtered_table)} rows, {len(filtered_table[0]) if filtered_table else 0} columns")
    print(f"Adjusted merged cells: {merged_cells}")
    return filtered_table, merged_cells